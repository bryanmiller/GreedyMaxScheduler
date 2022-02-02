#!/usr/bin/env python
# coding: utf-8

# Code for manipulating ODB Extractor json files and initial sequence atom definitions
# Bryan Miller
# 2021-11-24

import os
import sys
import json
import gzip
import numpy as np
import matplotlib.pyplot as plt
from scipy.signal import find_peaks
import requests

from openpyxl import Workbook
from openpyxl import load_workbook


fpuinst = {'GSAOI': 'instrument:utilityWheel', 'GPI': 'instrument:observingMode', 'Flamingos2': 'instrument:fpu',
           'NIFS': 'instrument:mask', 'GNIRS': 'instrument:slitWidth', 'GMOS-N': 'instrument:fpu',
           'GMOS-S': 'instrument:fpu', 'NIRI': 'instrument:mask'}

gpi_filter_wav = {'Y': 1.05, 'J': 1.25, 'H': 1.65, 'K1': 2.05, 'K2': 2.25}
nifs_filter_wav = {'ZJ': 1.05, 'JH': 1.25, 'HK': 2.20}


def find_filter(input, filter_dict):
    """Match input string with filter list (in dictionary)"""

    filter = ''
    filters = list(filter_dict.keys())
    for filt in filters:
        if filt in input:
            filter = filt
            break
    return filter


def uniquelist(seq):
    """Make a list of unique values"""
    # http://stackoverflow.com/questions/480214/how-do-you-remove-duplicates-from-a-list-in-python-whilst-preserving-order
    seen = set()
    seen_add = seen.add
    return [x for x in seq if not (x in seen or seen_add(x))]


def searchlist(val, alist):
    """Search for existence of val in any element of alist"""
    found = False
    for elem in alist:
        if val in elem:
            found = True
            break
    return found


def shortid(idin):
    """Return short form of obsid or data label"""
    idsp = idin.split('-')
    #         print(obsidsp)
    idout = idsp[0][1] + idsp[1][2:5] + '-' + idsp[2] + '-' + idsp[3] + '[' + idsp[4] + ']'
    if len(idsp) == 6:
        idout += '-' + idsp[5]
    return idout


def odb_json(progid, path='None', overwrite=False, verbose=False):
    """
    Download json of ODB program information

    Parameters
        progid:  Program ID of program to extract
        path:    Path for json files
        overwrite: Overwrite any existing json files?
        verbose: Verbose output?

    Return
        json_result:   JSON query result as a list of dictionaries
    """

    if progid == "":
        print('odb_json: program id not given.')
        raise ValueError('Program id not given.')

    file = progid + '.json.gz'
    if not overwrite and path != 'None' and os.path.exists(os.path.join(path, file)):
        with gzip.open(os.path.join(path, file), 'r') as fin:
            json_bytes = fin.read()

        json_str = json_bytes.decode('utf-8')
        json_result = json.loads(json_str)
    else:
        response = requests.get(
            'http://gnodbscheduler.hi.gemini.edu:8442/programexport?id=' + progid)
        try:
            response.raise_for_status()
        except requests.exceptions.HTTPError as exc:
            print('odb_json: request failed: {}'.format(response.text))
            raise exc
        else:
            json_result = response.json()
            if (overwrite or not os.path.exists(os.path.join(path, file))) and path != 'None':
                json_str = json.dumps(json_result, indent=2)
                with gzip.open(os.path.join(path, file), 'wb') as fout:
                    fout.write(json_str.encode('utf-8'))

        if verbose:
            print(response.url)
        # print(response.text)

    return json_result

# --------------


def obsmode(config):
    """Determine the observation mode (e.g. imaging, longslit, mos, ifu,..."""

    mode = 'unknown'
    if searchlist('GMOS', config['inst']):
        if 'MIRROR' in config['disperser']:
            mode = 'imaging'
        elif searchlist('arcsec', config['fpu']):
            mode = 'longslit'
        elif searchlist('IFU', config['fpu']):
            mode = 'ifu'
        elif 'CUSTOM_MASK' in config['fpu']:
            mode = 'mos'
    elif config['inst'] in ["GSAOI", "'Alopeke", "Zorro"]:
        mode = 'imaging'
    elif config['inst'] in ['IGRINS', 'MAROON-X']:
        mode = 'longslit'
    elif config['inst'] in ['GHOST', 'MAROON-X', 'GRACES', 'Phoenix']:
        mode = 'xd'
    elif config['inst'] == 'Flamingos2':
        if searchlist('LONGSLIT', config['fpu']):
            mode = 'longslit'
        if searchlist('FPU_NONE', config['fpu']) \
                and searchlist('IMAGING', config['disperser']):
            mode = 'imaging'
    elif config['inst'] == 'NIRI':
        if searchlist('NONE', config['disperser']) and searchlist('MASK_IMAGING', config['fpu']):
            mode = 'imaging'
    elif config['inst'] == 'NIFS':
        mode = 'ifu'
    elif config['inst'] == 'GNIRS':
        if searchlist('mirror', config['disperser']):
            mode = 'imaging'
        elif searchlist('XD', config['disperser']):
            mode = 'xd'
        else:
            mode = 'longslit'
    elif config['inst'] == 'GPI':
        if searchlist('CORON', config['fpu']):
            mode = 'coron'
        elif searchlist('NRM', config['fpu']):
            mode = 'nrm'
        elif searchlist('DIRECT', config['fpu']):
            mode = 'imaging'
        else:
            mode = 'ifu'

    return mode

# --------------


def guide_state(step):
    """Determine if guiding is on/off for a sequence step"""
    # One could also extract the guider used if needed
    guiding = False
    for key in list(step.keys()):
        if 'guideWith' in key:
            if step[key] == 'guide':
                guiding = True
                break
    return guiding

# --------------


def select_qastate(states):
    """Return the qastate based on precedence

        states: list of observe states from the ODB extractor obsLog
    """
    qastate = ''

    # Precedence order for observation classes.
    qastate_order = ['NONE', 'UNDEFINED', 'FAIL', 'USABLE', 'PASS']

    # Set the qastate for the entire observation based on precedence
    for state in qastate_order:
        if state in states:
            qastate = state
            break

    return qastate

# --------------


def select_obsclass(classes):
    """Return the obsclass based on precedence

        classes: list of observe classes from the ODB extractor
    """
    obsclass = ''

    # Precedence order for observation classes.
    obsclass_order = ['SCIENCE', 'PROGCAL', 'PARTNERCAL', 'ACQ', 'ACQCAL', 'DAYCAL']

    # Set the obsclass for the entire observation based on obsclass precedence
    for oclass in obsclass_order:
        if oclass in classes:
            obsclass = oclass
            break

    return obsclass

# --------------


def autocorr_lag(x, plot=False):
    '''Test for patterns with auto-correlation'''

    # Auto correlation
    result = np.correlate(x, x, mode='full')
    corrmax = np.max(result)
    if corrmax != 0.0:
        result = result / corrmax
    #     print(result)
    #     print(result.size, result.size//2)
    if plot:
        plt.plot(result[result.size // 2:])

    # Pattern offset using first prominent peak
    # lag = np.argmax(result[result.size//2 + 1:]) + 1
    peaks, prop = find_peaks(result[result.size // 2:], height=(0, None), prominence=(0.25, None))
    #     print(peaks)
    #     print(prop)
    lag = 0
    if len(peaks) > 0:
        lag = peaks[0]

    return lag

# --------------


def findatoms(observation, verbose=False, ws=None, fid=sys.stdout):
    """Analyze a json observing sequence from the ODB and define atoms."""

    classes = []
    guiding = []
    qastates = []
    atoms = []
    natom = 0

    # Make dictionary out of obslog to get QA state
    obslog = {}
    datalabels = []
    if 'obsLog' in observation.keys():
        for log_entry in observation['obsLog']:
            obslog[log_entry['label']] = {'qaState': log_entry['qaState'], 'filename': log_entry['filename']}
        datalabels = list(obslog.keys())

    # Sequence analysis
    sequence = observation['sequence']
    nsteps = len(sequence)

    # First pass to analyze offsets and exptimes/coadds
    do_not_split = False
    exptimes = []
    coadds = []
    qoffsets = []
    poffsets = []
    qoffsets_sky = []
    poffsets_sky = []
    config = {'inst': '', 'fpu': [], 'disperser': [], 'filter': [], 'wavelength': []}
    for ii, step in enumerate(sequence):
        step_keys = list(step.keys())

        inst = step['instrument:instrument']
        #     print(inst, fpuinst[inst])
        if inst == 'Visitor Instrument':
            inst = step['instrument:name'].split(' ')[0]
            if inst in ["'Alopeke", "Zorro"]:
                fpu = 'None'
            else:
                fpu = inst
        else:
            fpu = step[fpuinst[inst]]
        config['inst'] = inst
        config['fpu'].append(fpu)

        if 'instrument:disperser' in step_keys:
            disperser = step['instrument:disperser']
        elif inst in ['IGRINS', 'MAROON-X']:
            disperser = inst
        else:
            disperser = 'None'
        if inst == 'GNIRS':
            if step['instrument:acquisitionMirror'] == 'in' and \
                    step['instrument:decker'] == 'acquisition':
                disperser = 'mirror'
            else:
                disperser = disperser.strip('grating') + step['instrument:crossDispersed']
        elif inst == 'Flamingos2' and fpu == 'FPU_NONE':
            if step['instrument:decker'] == 'IMAGING':
                disperser = step['instrument:decker']
        config['disperser'].append(disperser)

        if 'instrument:filter' in step_keys:
            filter = step['instrument:filter']
        elif inst == 'GPI':
            filter = find_filter(fpu, gpi_filter_wav)
        else:
            if inst == 'GNIRS':
                filter = 'None'
            else:
                filter = 'Unknown'
        if inst == 'NIFS' and 'Same as Disperser' in filter:
            for filt in list(nifs_filter_wav.keys()):
                if disperser[0] in filt:
                    filter = filt
                    break
        config['filter'].append(filter)

        if inst == 'GPI':
            wavelength = gpi_filter_wav[filter]
        else:
            wavelength = float(step['instrument:observingWavelength'])
        config['wavelength'].append(wavelength)

        p = 0.0
        q = 0.0
        # Just want exposures on sky for dither pattern analysis
        if step['observe:observeType'].upper() not in ['FLAT', 'ARC', 'DARK', 'BIAS']:
            if 'telescope:p' in step_keys:
                p = float(step['telescope:p'])
            else:
                p = 0.0
            poffsets_sky.append(p)

            if 'telescope:q' in step_keys:
                q = float(step['telescope:q'])
            else:
                q = 0.0
            qoffsets_sky.append(q)

        if 'observe:coadds' in step_keys:
            ncoadds = int(step['observe:coadds'])
        else:
            ncoadds = 1
        poffsets.append(p)
        qoffsets.append(q)

        exptimes.append(float(step['observe:exposureTime']))
        coadds.append(ncoadds)

    mode = obsmode(config)
    if verbose:
        print(config)
        print(mode)

    if config['inst'] == 'GPI':
        do_not_split = True

    # Analyze sky offset patterns using auto-correlation
    # The lag is the length of any pattern, 0 means no repeating pattern
    plag = 0
    qlag = 0
    offset_lag = 0
    if do_not_split:
        offset_lag = nsteps
    else:
        if len(poffsets_sky) > 1:
            plag = autocorr_lag(np.array(poffsets_sky))
        if len(qoffsets_sky) > 1:
            qlag = autocorr_lag(np.array(qoffsets_sky))

        # Special cases
        if plag == 0 and qlag == 0 and len(qoffsets_sky) == 4:
            # single ABBA pattern, which the auto-correlation won't find
            if qoffsets_sky[0] == qoffsets_sky[3] and qoffsets_sky[1] == qoffsets_sky[2]:
                qlag = 4
        elif len(qoffsets_sky) == 2:
            # If only two steps, put them together, might be AB, also silly to split only two steps
            qlag = 2

        offset_lag = qlag
        if plag > 0 and plag != qlag:
            offset_lag = 0

        # Turn off offset groups if the length is longer than some limit
    #     if offset_lag > 0 and plag > 0 and qlag > 0 and \
    #        np.sum(np.asarray(exptimes[0:offset_lag]) * np.asarray(coadds[0:offset_lag])) > 600.:
    #         offset_lag = 0

    if verbose:
        print('Offset lags: ', plag, qlag, offset_lag, file=fid)

    # Group by changes in exptimes/coadds?
    exptime_groups = False
    #     if len(uniquelist(exptimes)) > 1 or len(uniquelist(coadds)) > 1:
    #         exptime_groups = True

    # Second pass to determine atom properties
    # Write results to the Excel worksheet if given
    if ws is not None:
        # Columns
        columns = ['datalab', 'class', 'type', 'inst', 'exec_time', 'exptime', 'coadds', 'fpu', 'filter',
                   'disperser', 'wavelength', 'p', 'q', 'guiding', 'qa_state', 'atom']
        row = 1
        for jj, col in enumerate(columns):
            _ = ws.cell(column=jj + 1, row=row, value="{0}".format(col))
        row += 1

    npattern = offset_lag
    noffsets = 0
    for ii, step in enumerate(sequence):
        nextatom = False

        datalab = step['observe:dataLabel']
        if datalab in datalabels:
            qastate = obslog[datalab]['qaState']
        else:
            qastate = 'NONE'

        observe_class = step['observe:class']

        step_time = step['totalTime'] / 1000.

        atomstr = 'Atom for: '
        # Any wavelength/filter change is a new atom
        if ii == 0 or (ii > 0 and config['wavelength'][ii] != config['wavelength'][ii - 1]):
            nextatom = True
            atomstr += 'wavelength, '

        # A change in exposure time or coadds is a new atom for science exposures
        if step['observe:observeType'].upper() not in ['FLAT', 'ARC', 'DARK', 'BIAS']:
            if observe_class.upper() == 'SCIENCE' and \
                    ii > 0 and (exptimes[ii] != exptimes[ii_prevobj] or coadds[ii] != coadds[ii_prevobj]):
                nextatom = True
                atomstr += 'exposure time change, '

            # Offsets - a new offset pattern is a new atom
            #         print('npattern: ', npattern)
            if not (offset_lag == 0 and exptime_groups == True):
                # For NIR imaging, need to have at least two offset positions if no repeating pattern
                # New atom after every 2nd offset (noffsets is odd)
                if mode == 'imaging' and offset_lag == 0 and all([w > 1.0 for w in config['wavelength']]):
                    if ii == 0:
                        noffsets += 1
                    else:
                        if poffsets[ii] != poffsets[ii_prevobj] or qoffsets[ii] != qoffsets[ii_prevobj]:
                            noffsets += 1
                    if noffsets % 2 == 1:
                        nextatom = True
                        atomstr += 'offset pattern'
                else:
                    npattern -= 1
                    if npattern < 0:
                        nextatom = True
                        atomstr += 'offset pattern'
                        npattern = offset_lag - 1

            ii_prevobj = ii

        # New atom?
        if nextatom:
            # Get class, qastate, guiding for previous atom
            if natom > 0:
                atoms[-1]['qa_state'] = select_qastate(qastates)
                if atoms[-1]['qa_state'] != 'NONE':
                    atoms[-1]['observed'] = True
                atoms[-1]['class'] = select_obsclass(classes)
                if verbose:
                    print('QA states: ', qastates, file=fid)
                    print('Classes: ', classes, file=fid)
                    print('Guiding: ', guiding, file=fid)
                atoms[-1]['guide_state'] = any(guiding)
                atoms[-1]['wavelength'] = config['wavelength'][ii]
                atoms[-1]['required_resources']['inst'] = config['inst']
                atoms[-1]['required_resources']['filter'] = config['filter'][ii]
                atoms[-1]['required_resources']['disperser'] = config['disperser'][ii]
                atoms[-1]['required_resources']['fpu'] = config['fpu'][ii]

                # Print basic atom info
                print(" \t exec_time: {:7.2f}, prog_time: {:7.2f}, part_time: {:7.2f}, guide_state: {}".
                      format(atoms[-1]['exec_time'], atoms[-1]['prog_time'], atoms[-1]['part_time'],
                             atoms[-1]['guide_state']), file=fid)

            # New atom entry
            print(atomstr, file=fid)
            natom += 1
            atoms.append({'id': natom, 'exec_time': 0.0, 'prog_time': 0.0, 'part_time': 0.0,
                          'class': 'NONE', 'observed': False, 'qa_state': 'NONE', 'guide_state': False,
                          'wavelength': 0.0,
                          'required_resources': {'inst': 'NONE', 'filter': 'NONE', 'disperser': 'NONE',
                                                 'fpu': 'NONE'}})
            classes = []
            guiding = []
            qastates = []
            if step['observe:observeType'].upper() in ['FLAT', 'ARC', 'DARK', 'BIAS'] \
                    and npattern == 0:
                npattern = offset_lag
            noffsets = 1

        # Update atom
        qastates.append(qastate.upper())
        classes.append(observe_class.upper())
        guiding.append(guide_state(step))

        atoms[-1]['exec_time'] += step_time

        atomlabel = natom
        if 'partnerCal' in observe_class:
            atomlabel *= 10
            atoms[-1]['part_time'] += step_time
        else:
            atoms[-1]['prog_time'] += step_time

        print('{:17} {:3} {:7.2f} {:3d} {:10} {:12} {:10} {:12} {:6.4f} {:8.2f} {:8.2f} {:1} {:3d}'.format(
            shortid(datalab),
            observe_class[0:3], exptimes[ii], coadds[ii], config['inst'], config['fpu'][ii],
            config['filter'][ii], config['disperser'][ii], config['wavelength'][ii], poffsets[ii],
            qoffsets[ii], guiding[-1], atomlabel),
              file=fid)
        if ws is not None:
            # Columns
            #             columns = ['datalab', 'class', 'type', inst', 'exec_time', 'exptime', 'coadds', 'fpu', 'filter',
            #                        'disperser', 'wavelength', 'p', 'q', 'guiding', 'qa_state', 'atom']
            data = [datalab, observe_class.upper(), step['observe:observeType'].upper(), config['inst'],
                    step_time, exptimes[ii], coadds[ii], config['fpu'][ii],
                    config['filter'][ii], config['disperser'][ii], config['wavelength'][ii], poffsets[ii],
                    qoffsets[ii], guiding[-1], qastate.upper(), atomlabel]
            for jj in range(len(columns)):
                _ = ws.cell(column=jj + 1, row=row, value=data[jj])
            row += 1

    #     print(atoms)
    # Get class/state for last atom
    if natom > 0:
        atoms[-1]['qa_state'] = select_qastate(qastates)
        if atoms[-1]['qa_state'] != 'NONE':
            atoms[-1]['observed'] = True
        atoms[-1]['class'] = select_obsclass(classes)
        if verbose:
            print('QA states: ', qastates, file=fid)
            print('Classes: ', classes, file=fid)
            print('Guiding: ', guiding, file=fid)
        atoms[-1]['guide_state'] = any(guiding)
        atoms[-1]['wavelength'] = config['wavelength'][ii]
        atoms[-1]['required_resources']['inst'] = config['inst']
        atoms[-1]['required_resources']['filter'] = config['filter'][ii]
        atoms[-1]['required_resources']['disperser'] = config['disperser'][ii]
        atoms[-1]['required_resources']['fpu'] = config['fpu'][ii]

        # Print basic atom info
        print(" \t exec_time: {:7.2f}, prog_time: {:7.2f}, part_time: {:7.2f}, guide_state: {}".
              format(atoms[-1]['exec_time'], atoms[-1]['prog_time'], atoms[-1]['part_time'],
                     atoms[-1]['guide_state']), file=fid)

    return atoms

# --------------


def group_proc(group, sel_obs_class=['SCIENCE', 'PROGCAL', 'PARTNERCAL', 'ACQ', 'ACQCAL', 'DAYCAL'],
               sel_obs_status=['PHASE_2', 'FOR_REVIEW', 'IN_REVIEW', 'FOR_ACTIVATION', 'ON_HOLD', 'READY',
                               'ONGOING', 'OBSERVED', 'INACTIVE'], fid=sys.stdout, wb=None, verbose=False):
    """Process observations within groups"""

    ws = None
    obsnum = []
    for item in list(group.keys()):
        obsid = ''
        if 'OBSERVATION' in item:
            obsid = group[item]['observationId']
            obsnum.append(int(item.split('-')[1]))


    if len(obsnum) > 0:
        isrt = np.argsort(obsnum)
        for ii in isrt:
            obs_program_used = 0.0
            obs_partner_used = 0.0
            item = 'OBSERVATION_BASIC-' + str(obsnum[ii])
            #     obsid = program[prog][group][item]['sequence'][0]['ocs:observationId']
            obsid = group[item]['observationId']
            print(f"{obsnum[ii], obsid}", file=fid)
            obs_class = group[item]['obsClass'].upper()
            phase2stat = group[item]['phase2Status'].upper()
            obs_stat = group[item]['obsStatus'].upper()
            #             print(obs_class, phase2stat, obs_stat)
            if obs_class in sel_obs_class and obs_stat in sel_obs_status:
                if wb is not None:
                    ws = wb.create_sheet(title=obsid.split('-')[-1])
                #                     ws['A1'] = obsid

                # Atoms in each sequence
                atoms = findatoms(group[item], verbose=verbose, ws=ws, fid=fid)
                # Summary of atoms
                #                 classes = []
                #                 qastates = []
                #                 for atom in atoms:
                #                     print('Atom ', atom['id'])
                #                     for key in atom.keys():
                #                         print(f" \t {key}: {atom[key]}")
                #                         if key == 'class':
                #                             classes.append(atom[key])
                #                         if key == 'qa_state':
                #                             qastates.append(atom[key])
                #                             if atom[key].upper() == 'PASS':
                #                                 obs_program_used += atom['prog_time']
                #                                 obs_partner_used += atom['part_time']
                #                 obsclass = select_obsclass(classes)
                #                 print(f" Obsclass: {obsclass}")
                #                 obs_qastate = select_qastate(qastates)
                #                 print(f" QAstate (atoms): {obs_qastate}")
                #                 print(f" qaState (ODB): {group[item]['qaState']}")
                #                 if group[item]['qaState'].upper() == 'PASS':
                #                     if group[item]['obsClass'] in ['science', 'progCal']:
                #                         obs_program_used += float(group[item]['setupTime'])/1000.
                #                     elif group[item]['obsClass'] in ['partnerCal']:
                #                         obs_partner_used += float(group[item]['setupTime'])/1000.

                #                 print(f" program_used: {obs_program_used}")
                #                 print(f" partner_used: {obs_partner_used}")
                print('', file=fid)

            print('', file=fid)

    return


def prog_proc(program, sel_obs_class=['SCIENCE', 'PROGCAL', 'PARTNERCAL', 'ACQ', 'ACQCAL', 'DAYCAL'], \
              sel_obs_status=['PHASE_2', 'FOR_REVIEW', 'IN_REVIEW', 'FOR_ACTIVATION', 'ON_HOLD', 'READY', \
                              'ONGOING', 'OBSERVED', 'INACTIVE'], fid=sys.stdout, xls=None, verbose=False):
    """Process top-level of program"""

    wb = None
    grpnum = []
    grplist = []
    for prog in list(program.keys()):
        #     print(list(program[prog].keys()))
        print(f"**** {program[prog]['programId']} ****", file=fid)

        # Excel output?
        if xls is not None:
            wb = Workbook()
            ws = wb.active

        # First pass to count and record groups, also show the ATOM note
        for item in list(program[prog].keys()):
            if 'INFO' in item:
                if 'ATOM' in program[prog][item]['title']:
                    print(f"{program[prog][item]['title']}: {program[prog][item]['text']}\n", file=fid)
                    # Comment
                    if xls is not None:
                        ws['A1'] = program[prog]['programId']
                        ws['A2'] = 'ATOMS'
                        ws['B2'] = program[prog][item]['text']
            if 'GROUP' in item:
                #                 print(item, program[prog][item]['name'])
                #         print(program[prog][item])
                grpnum.append(int(item.split('-')[1]))
                grplist.append(item.split('-')[0])

        if len(grpnum) > 0:
            # Second pass to put the groups in the same order as in the program
            #             print(grplist)
            #             print(grpnum)
            isrt = np.argsort(grpnum)
            #     print(grpnum)
            for ii in isrt:
                group = grplist[ii] + '-' + str(grpnum[ii])
                print(group, program[prog][group]['name'], file=fid)
                group_proc(program[prog][group], sel_obs_class=sel_obs_class, sel_obs_status=sel_obs_status,
                           fid=fid, wb=wb, verbose=verbose)

        # Process any observations at the root level
        group_proc(program[prog], sel_obs_class=sel_obs_class, sel_obs_status=sel_obs_status, fid=fid,
                   wb=wb, verbose=verbose)
        print('', file=fid)

    if xls is not None:
        wb.save(filename=xls)

    return


def printseq(sequence, comment='', csv=False, path=''):
    '''Print basic configuration information about a sequence, with an option to write to a csv file'''

    atom = '1'
    if csv and path != '':
        obsid = sequence[0]['ocs:observationId']
        filename = os.path.join(path, obsid + '_seq.csv')
        f = open(filename, 'w')
        print('{},{}'.format('comment', comment), file=f)
        print('{},{},{},{},{},{},{},{},{},{},{},{}'.format('datalab', 'class', 'exptime', 'coadds', 'inst', 'fpu',
                                                        'filter_name', 'disperser', 'wavelength', 'p', 'q', 'atom'), file=f)

    for step in list(sequence):
        datalab = step['observe:dataLabel']
        observe_class = step['observe:class']
        exptime = step['observe:exposureTime']
        inst = step['instrument:instrument']
    #     print(inst, fpuinst[inst])
        fpu = step[fpuinst[inst]]
        if 'instrument:filter_name' in step.keys():
            filter_name = step["instrument:filter_name"]
        else:
            filter_name = 'None'
        wavelength = step['instrument:observingWavelength']
        if 'GMOS' in inst:
            coadds = '1'
            # convert wavelength to microns
#             wavelength = '{:5.3f}'.format(float(wavelength) / 1000.)
        else:
            coadds = step['observe:coadds']
        disperser = step['instrument:disperser']
        if 'telescope:p' in step.keys():
            p = step['telescope:p']
        else:
            p = '0.0'
        if 'telescope:q' in step.keys():
            q = step['telescope:q']
        else:
            q = '0.0'    
        print('{:25} {:10} {:7} {:3} {:10} {:20} {:12} {:12} {:7} {:5} {:5}'.format(datalab, observe_class, exptime, coadds,
                                                                       inst, fpu, filter_name, disperser, wavelength, p, q))
        if csv and path != '':
            print('{},{},{},{},{},{},{},{},{},{},{},{}'.format(datalab, observe_class, exptime, coadds, inst, fpu,
                                                               filter_name, disperser, wavelength, p, q, atom), file=f)

    if csv and path != '':
        f.close()


def seqxlsx(sequence, comment='', path=''):
    '''Write sequence information to an Excel spreadsheet'''

    obsid = sequence[0]['ocs:observationId']
    filename = os.path.join(path, obsid + '_seq.xlsx')
    wb = Workbook()
    ws = wb.active
    
    atom = '1'
    
    # Comment
    ws['A1'] = 'comment'
    ws['B1'] = comment

    # Columns
    columns = ['datalab', 'class', 'inst', 'exptime', 'coadds', 'fpu', 'filter_name',
               'disperser', 'wavelength', 'p', 'q', 'atom']
    
    row = 2
    for ii, col in enumerate(columns):
        _ = ws.cell(column=ii+1, row=row, value="{0}".format(col))
    row += 1
        
#     print('{},{}'.format('comment', comment), file=f)
#     print('{},{},{},{},{},{},{},{},{},{},{}'.format('datalab', 'class', 'exptime', 'coadds', 'inst', 'fpu', 
#                                                            'disperser', 'wavelength', 'p', 'q', 'atom'), file=f)

    for step in list(sequence):
        data = []
        data.append(step['observe:dataLabel'])
        data.append(step['observe:class'])
        inst = step['instrument:instrument']
        data.append(inst)
    #     print(inst, fpuinst[inst])
        data.append(float(step['observe:exposureTime']))
        if 'GMOS' in inst:
            coadds = '1'
            # convert wavelength to microns
#             wavelength = '{:5.3f}'.format(float(wavelength) / 1000.)
        else:
            coadds = step['observe:coadds']
        data.append(int(coadds))
        data.append(step[fpuinst[inst]])
        if 'instrument:filter_name' in step.keys():
            filter_name = step["instrument:filter_name"]
        else:
            filter_name = 'None'
        data.append(filter_name)
        data.append(step['instrument:disperser'])
        data.append(float(step['instrument:observingWavelength']))
        if 'telescope:p' in step.keys():
            p = step['telescope:p']
        else:
            p = '0.0'
        data.append(float(p))
        if 'telescope:q' in step.keys():
            q = step['telescope:q']
        else:
            q = '0.0'  
        data.append(float(q))
        data.append(int(atom))
        print(data)
        
        for ii in range(len(columns)):
            _ = ws.cell(column=ii+1, row=row, value=data[ii])
        row += 1
    
    wb.save(filename)
    return


def readseq(file, path):
    '''Read sequence information from a csv file'''

    filename = os.path.join(path, file)
    f = open(filename, 'r')
    
    sequence = {}
    
    # Read and parse csv file: first line is a comment, second has column headings
    nline = 0
    for line in f:
#         line = line.rstrip('\n')
        values = line.rstrip('\n').split(',')
        if nline == 0:
            sequence['comment'] = values[1]
        elif nline == 1:
            columns = list(values)
            print(columns)
            for col in columns:
                sequence[col.strip(' ')] = []
        else:
            for i, val in enumerate(values):
                sequence[columns[i].strip(' ')].append(val.strip(' '))
        nline += 1
        
    f.close()
    
    return sequence


def xlsxseq(file, path):
    '''Read sequence information from an Excel spreadsheet'''

    filename = os.path.join(path, file)
    
    wb = load_workbook(filename=filename)
    ws = wb.active
    
    sequence = {}
    
    row = 1
    sequence['comment'] = ws.cell(column=2, row=row).value
    row += 1
    
    columns = []
    # Eventually ready the number of columns in the sheet
    for ii in range(26):
        col = ws.cell(column=ii+1, row=row).value
        if col is not None:
            columns.append(col)
            sequence[col] = []
        else:
            break
    row += 1
#     print(columns)

    while ws.cell(column=1, row=row).value is not None:
        for jj, col in enumerate(columns):
            sequence[col].append(ws.cell(column=jj+1, row=row).value)
        row += 1
    
    return sequence


def xlsxatoms(file, path, sheet='None', verbose=False):
    # Read a spreadsheet created by findatoms

    filename = os.path.join(path, file)

    atoms_dict = {}

    wb = load_workbook(filename=filename)

    sheets = []
    if sheet == 'None':
        # Read all sheets except the first
        sheets = wb.sheetnames
        sheets.remove('Sheet')
    elif sheet != 'None' and sheet in wb.sheetnames:
        sheets = [sheet]
    else:
        print(f"Sheet {sheet} not found.")

    for sheet in sheets:
        ws = wb[sheet]
        print(f"Sheet {ws.title}")
        # Columns
        #         columns = ['datalab', 'class', 'type', inst', 'exec_time', 'exptime', 'coadds', 'fpu', 'filter',
        #                    'disperser', 'wavelength', 'p', 'q', 'guiding', 'qa_state', 'atom']
        sequence = {}
        columns = []
        # Eventually read the number of columns in the sheet
        row = 1
        for ii in range(16):
            col = ws.cell(column=ii + 1, row=row).value
            if col is not None:
                columns.append(col)
                sequence[col] = []
            else:
                break
        row += 1

        natom = 0
        atoms = []
        classes = []
        guiding = []
        qastates = []
        while ws.cell(column=1, row=row).value is not None:
            nextatom = False
            for jj, col in enumerate(columns):
                sequence[col].append(ws.cell(column=jj + 1, row=row).value)
            #             print(sequence['datalab'][-1], sequence['inst'][-1], sequence['atom'][-1])

            # obsid
            if natom == 0:
                datalab = sequence['datalab'][-1]
                obsid = datalab[0:datalab.rfind('-')]

            if natom == 0 or (natom > 0 and sequence['atom'][-1] != sequence['atom'][-2]):
                nextatom = True

            # New atom?
            if nextatom:
                # Get class, qastate, guiding for previous atom
                if natom > 0:
                    atoms[-1]['qa_state'] = select_qastate(qastates)
                    if atoms[-1]['qa_state'] != 'NONE':
                        atoms[-1]['observed'] = True
                    atoms[-1]['class'] = select_obsclass(classes)
                    if verbose:
                        print('QA states: ', qastates)
                        print('Classes: ', classes)
                        print('Guiding: ', guiding)
                    atoms[-1]['guide_state'] = any(guiding)
                    atoms[-1]['wavelength'] = sequence['wavelength'][-1]
                    atoms[-1]['required_resources']['inst'] = sequence['inst'][-1]
                    atoms[-1]['required_resources']['filter'] = sequence['filter'][-1]
                    atoms[-1]['required_resources']['disperser'] = sequence['disperser'][-1]
                    atoms[-1]['required_resources']['fpu'] = sequence['fpu'][-1]

                    # Print basic atom info
                    print(" \t exec_time: {:7.2f}, prog_time: {:7.2f}, part_time: {:7.2f}, guide_state: {}".
                          format(atoms[-1]['exec_time'], atoms[-1]['prog_time'], atoms[-1]['part_time'],
                                 atoms[-1]['guide_state']))
                natom += 1
                atoms.append({'id': natom, 'exec_time': 0.0, 'prog_time': 0.0, 'part_time': 0.0,
                              'class': 'NONE', 'observed': False, 'qa_state': 'NONE', 'guide_state': False,
                              'wavelength': 0.0,
                              'required_resources': {'inst': 'NONE', 'filter': 'NONE', 'disperser': 'NONE',
                                                     'fpu': 'NONE'}})
                classes = []
                guiding = []
                qastates = []

            qastates.append(sequence['qa_state'][-1])
            classes.append(sequence['class'][-1])
            guiding.append(sequence['guiding'][-1])

            atoms[-1]['exec_time'] += sequence['exec_time'][-1]

            atomlabel = natom
            if 'PARTNERCAL' in sequence['class'][-1]:
                atomlabel *= 10
                atoms[-1]['part_time'] += sequence['exec_time'][-1]
            else:
                atoms[-1]['prog_time'] += sequence['exec_time'][-1]

            row += 1

        # Get class/state for last atom
        if natom > 0:
            atoms[-1]['qa_state'] = select_qastate(qastates)
            if atoms[-1]['qa_state'] != 'NONE':
                atoms[-1]['observed'] = True
            atoms[-1]['class'] = select_obsclass(classes)
            if verbose:
                print('QA states: ', qastates)
                print('Classes: ', classes)
                print('Guiding: ', guiding)
            atoms[-1]['guide_state'] = any(guiding)
            atoms[-1]['wavelength'] = sequence['wavelength'][-1]
            atoms[-1]['required_resources']['inst'] = sequence['inst'][-1]
            atoms[-1]['required_resources']['filter'] = sequence['filter'][-1]
            atoms[-1]['required_resources']['disperser'] = sequence['disperser'][-1]
            atoms[-1]['required_resources']['fpu'] = sequence['fpu'][-1]

            # Print basic atom info
            print(" \t exec_time: {:7.2f}, prog_time: {:7.2f}, part_time: {:7.2f}, guide_state: {}". \
                  format(atoms[-1]['exec_time'], atoms[-1]['prog_time'], atoms[-1]['part_time'], \
                         atoms[-1]['guide_state']))

        atoms_dict[obsid] = atoms

    wb.close()

    return atoms_dict


if __name__ == '__main__':

    path = './data'
    print(path)

    # programs = ['GN-2018B-Q-101', 'GN-2018B-Q-106', 'GN-2018B-FT-206', ]
    # programs = ['GN-2018B-DD-104', 'GS-2018B-Q-226']
    programs = ['GS-2018B-Q-102']
    f = sys.stdout
    # f = open(os.path.join(path, 'atoms_test.txt'), 'w')
    for progid in programs:
        xlsout = os.path.join(path, progid + '.xlsx')
        program = odb_json(progid, path=path)
        prog_proc(program, sel_obs_class=['SCIENCE', 'PROGCAL', 'PARTNERCAL'],
                    sel_obs_status=['READY', 'ONGOING', 'OBSERVED'], verbose=False, xls=xlsout, fid=f)
    if f != sys.stdout:
        f.close()

    atoms = xlsxatoms('GS-2018B-Q-102.xlsx', path, '23')
    print(atoms)
